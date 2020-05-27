import os
from openpyxl import *

def delete_col(filename):
    black_list = ['检查标准来源', '行业小类', '要求内容', '操作指南', '发现时间', '发现人姓名']
    wb = load_workbook(filename)

    ws = wb.active
    col_nums = ws.max_column  # 总列数
    list=[]
    for x in range(1, col_nums + 1):  # 循环所有列
        title = ws.cell(row=1, column=x).value  # 提取列名的
        for keyword in black_list:  # 判断是不是在要删除的里面
            # print(str(title))
            if keyword == str(title).strip():
                list.append(x)
    list.sort(reverse=True)
    for i in list:
        ws.delete_cols(i)
    wb.save(filename)

def delete_row(filename):
    black_list = ['检查新建用户的home目录的缺省访问权限', '修改vsftp回显信息', '禁止匿名FTP', '不存在无用用户组', '/etc/services文件权限', 'VSFTP用户权限名单', '检查是否指定用户组成员使用su命令', '删除或锁定无用账号', '分配不同账号，避免不同用户间共享账号', '禁用Telnet服务', 'Vsftp的chroot list配置', '配置记录cron行为日志功能', '记录系统日志及应用日志', '配置远程日志服务器', '日志文件读写权限', '关闭不必要的服务', '系统umask设置', 'passwd shadow group文件安全性配置', '检查密码长度及复杂度策略', '定时账户自动登出']
    wb = load_workbook(filename)
    ws = wb.active
    row_nums = ws.max_row  # 总行数
    list=[]
    for x in range(2, row_nums + 1):
        title = ws.cell(column=6, row=x).value #提取第六列数据
        keyword = str(title).strip()
        if keyword in black_list:
            pass
        else:
            list.append(x)
    list.sort(reverse=True)
    # map(ws.delete_rows, list)
    for i in list:
        ws.delete_rows(i)
    wb.save(filename) 

def readFiles(path):
    files = os.listdir(path)
    for file in files:
        if ".DS_Store" in file:
            continue
        execls = os.listdir(path+"/"+file)
        for execl in execls:
            if ".xlsx" in execl:
                delete_col(path+"/"+file+"/"+execl)
                delete_row(path+"/"+file+"/"+execl)
            else:
                continue

def readFile(path):
    files = os.listdir(path)
    for file in files:
        if ".docx" in file:
            continue
        delete_col(path+"/"+file)
        delete_row(path+"/"+file)
if __name__ == "__main__":
    readFiles('./aaa')
    # readFile('#000001-CRM')